import pdfplumber
import re
import sys
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

excelInfo = []

excelInfoSinClasi = []


# El diccionario sera un arreglo de tuplas que tendra costo y quien lo gasto
diccionarioTotales = {}

diccionarioHabilitados = {
    "Claudia": True,
    "Omar": False,
    "Hernes": False,
    "Pri": False,
    "PV": True,
    "Chihuahua": False,
    "QuintanaRoo": True,
    "CatyMonreal": False,
}

palabrasClavePri = ["#OTROS PRI", "PRI"]


palablasClaveClaudia = [
    "AGUA",
    "AMBIENTE",
    "BECAS",
    "DISCAPACIDAD",
    "FISCALIA",
    "MEDICAMENTO",
    "SEGURIDAD",
    "SUELO PROTEGIDO VITAMINA",
    "GASES AGUITA",
    "GASES VITAMINA",
    "PLANTA SOLAR FREE",
    "PLANTA SOLAR VITAMINA",
    "RETOVERDE",
    "MEDICINA",
    "MEDICINAS",
    "FICALIA",
    "SHEINBAUM",
    "SHEIMBAUM",
]

palablasClaveHernes = ["HERNES"]

pablabrasClaveOmar = ["OMAR", "HARFUCH"]

pablbrasClavePV = [
    "#PVEM",
    "#PVME",
    "Ambiental" "Videovigilancia",
    "Estres Laboral",
    "Cancel Infantil",
    "Pluvial",
    "Agua",
    "Animales Compañía",
    "Electrios",
    "Autos Electricos 3",
    "Superbowl",
    "Parque Vehicular",
    "Cero Emisiones",
    "Delfine",
    "Justicia",
    "Delitos",
    "Basura Organica",
    "DELFINES SPOT",
    "JUSTICIA",
    "PERMISOS",
    "ANIMALES DE COMPAÑÍA",
    "Banner ADOPCION",
    "CANCER INF",
    "Banner HUERTOS",
    "ESTRES LABORAL",
    "HUERTOS",
    "PERMISOS",
    "TALA ILEGAL",
    "Clinica",
    "Clinicas Macotas",
    "Permiso Laboral",
    "Huertos",
    "Permiso Padres",
]

pablbrasClaveChihuahua = ["#CH"]

pablbrasClaveQuintanaRoo = ["#QR", "GINO"]

pablbrasClaveCatyMonreal = ["CATY MONREAL"]


def sumarTotal(titulo, costo):
    if isCostoClaudia(titulo):
        diccionarioTotales["Claudia"] = diccionarioTotales.get("Claudia", 0) + costo
    elif isCostoOmar(titulo):
        diccionarioTotales["Omar"] = diccionarioTotales.get("Omar", 0) + costo
    elif isCostoPV(titulo):
        diccionarioTotales["PV"] = diccionarioTotales.get("PV", 0) + costo
    elif isCostoChihuahua(titulo):
        diccionarioTotales["Chihuahua"] = diccionarioTotales.get("Chihuahua", 0) + costo
    elif isCostoQuintanaRoo(titulo):
        diccionarioTotales["QuintanaRoo"] = (
            diccionarioTotales.get("QuintanaRoo", 0) + costo
        )
    elif isCostoCatyMonreal(titulo):
        diccionarioTotales["CatyMonreal"] = (
            diccionarioTotales.get("CatyMonreal", 0) + costo
        )
    elif isCostoHernes(titulo):
        diccionarioTotales["Hernes"] = diccionarioTotales.get("Hernes", 0) + costo
    elif isCostoPri(titulo):
        diccionarioTotales["Pri"] = diccionarioTotales.get("Pri", 0) + costo
    else:
        print("Movimiento no identificado", titulo, costo)
        excelInfoSinClasi.append([costo, titulo])
        diccionarioTotales["Otro"] = diccionarioTotales.get("Otro", 0) + costo


def isCostoPri(titulo):
    if not diccionarioHabilitados["Pri"]:
        return False
    for palabra in palabrasClavePri:
        if titulo.upper().__contains__(palabra):
            return True
    return False


def isCostoClaudia(titulo):
    if not diccionarioHabilitados["Claudia"]:
        return False

    # check if the title is in the list of claudia
    if titulo.upper().__contains__("CLAUDIA"):
        return True

    for palabra in palablasClaveClaudia:
        if titulo.upper().__contains__(palabra):
            return True
    return False


def isCostoOmar(titulo):
    if not diccionarioHabilitados["Omar"]:
        return False
    if titulo.upper().__contains__("OMAR"):
        return True
    else:
        return False


def isCostoHernes(titulo):
    if not diccionarioHabilitados["Hernes"]:
        return False
    if titulo.upper().__contains__("HERNES"):
        return True
    for palabra in palablasClaveHernes:
        if titulo.upper().__contains__(palabra):
            return True
    return False


def isCostoPV(titulo):
    if not diccionarioHabilitados["PV"]:
        return False
    for palabra in pablbrasClavePV:
        if titulo.upper().__contains__(palabra):
            return True
    return False


def isCostoChihuahua(titulo):
    if not diccionarioHabilitados["Chihuahua"]:
        return False
    for palabra in pablbrasClaveChihuahua:
        if titulo.upper().__contains__(palabra):
            return True
    return False


def isCostoQuintanaRoo(titulo):
    if not diccionarioHabilitados["QuintanaRoo"]:
        return False
    for palabra in pablbrasClaveQuintanaRoo:
        if titulo.upper().__contains__(palabra):
            return True
    return False


def isCostoCatyMonreal(titulo):
    if not diccionarioHabilitados["CatyMonreal"]:
        return False
    for palabra in pablbrasClaveCatyMonreal:
        if titulo.upper().__contains__(palabra):
            return True
    return False


def extraerDatos(pdf_path, excel_path):
    fechaArchivo = ""
    celdaTotlaes = 9
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            for line in texto.split("\n"):
                if line.__contains__("Fecha de la nota de pago pendiente/pago"):
                    # Get the next line
                    fechaArchivo = (
                        texto.split("\n")[texto.split("\n").index(line) + 1]
                        .split(",")[0]
                        .replace(".", "")
                    )

                if re.match(r"^#", line) and line.__contains__("Impresiones"):
                    elementos = line.split(" ")
                    elementos.remove("Impresiones")
                    costo = elementos[-1]
                    elementos.pop()
                    elementos.pop()
                    titulo = " ".join(elementos)
                    excelInfo.append([costo.replace("$", "").replace(",", "."), titulo])
                    # 1,054.43 -> 1.054.43 -> 1054.43
                    # remover todos los puntos excepto el ultimo
                    costoArray = (
                        costo.replace("$", "").replace(",", ".").split(".")
                    )  # = [1,054,43]
                    if len(costoArray) > 2:
                        decimales = costoArray.pop()
                        total = "".join(costoArray) + "." + decimales
                        sumarTotal(titulo, float(total))
                    else:
                        sumarTotal(
                            titulo, float(costo.replace("$", "").replace(",", "."))
                        )

    wb = Workbook()
    wb.create_sheet(fechaArchivo + "-facturas")

    # select the create sheet
    ws = wb[fechaArchivo + "-facturas"]
    ws.append(["Costo", "Titulo"])
    for row in excelInfo:
        ws.append(row)
    table = Table(
        displayName="Table" + fechaArchivo.replace(" ", "-"),
        ref="A0:B" + str(len(excelInfo) + 1),
    )
    ws.add_table(table)
    # Seleccionar ws anterior como activa

    # Agregar los elementos sin clasificar
    if len(excelInfoSinClasi) > 0:
        ws2 = wb.create_sheet("facturas-sin-clasificar")
        # select the create sheet
        ws2 = wb["facturas-sin-clasificar"]
        ws2.append(["Costo", "Titulo"])
        for row in excelInfoSinClasi:
            ws2.append(row)
        table = Table(
            displayName="Table-no-clasi",
            ref="A0:B" + str(len(excelInfoSinClasi) + 1),
        )
        ws2.add_table(table)

    for key in diccionarioTotales:
        cTitulo = ws["M" + str(celdaTotlaes)]
        cTitulo.value = "Total " + key
        cCosto = ws["N" + str(celdaTotlaes)]
        cCosto.value = diccionarioTotales[key]
        celdaTotlaes += 1

    cTotalTitle = ws["P9"]
    cTotalTitle.value = "Total"
    cTotal = ws["Q9"]
    cTotal.value = sum(diccionarioTotales.values())

    wb.save(excel_path + "facturas" + ".xlsx")
    wb.close()


def main():
    if len(sys.argv) != 2:
        print("Usage: python main.py <directory>")
        sys.exit(1)

    directorio = sys.argv[1]

    if not os.path.exists(directorio):
        print("The directory does not exist")
        sys.exit(1)
    else:
        index = 0
        for path in Path(directorio).rglob("*.pdf"):
            index += 1
            # get path delting the file name
            prevPath = str(path).replace(path.name, "")
            print("Procesando: " + str(path))
            extraerDatos(str(path), prevPath)

    sys.exit(0)


if __name__ == "__main__":
    main()
